"""Parser de planilha ERP (CSV/XLSX) — puro e testável.

A planilha é o segundo pilar do core ("subir XMLs E planilhas"): destrava
auditoria para qualquer construtora sem depender de API de ERP. Aqui só fazemos
parsing + auto-detecção de colunas; a carga ao canônico fica em load.py.

Auto-detecção: normaliza o cabeçalho (sem acento, minúsculo) e casa contra
sinônimos comuns de planilha de contas a pagar / lançamentos. Sem efeitos
colaterais — devolve linhas canônicas + o mapeamento detectado (útil p/ UX).
"""

from __future__ import annotations

import csv
import io
import unicodedata
from decimal import Decimal, InvalidOperation
from typing import Any

# campo canônico -> sinônimos de cabeçalho (já normalizados: sem acento, minúsculo)
_SYNONYMS: dict[str, set[str]] = {
    "fornecedor": {"fornecedor", "credor", "razao social", "nome fornecedor",
                   "beneficiario", "favorecido", "nome do fornecedor"},
    "cnpj": {"cnpj", "cnpj/cpf", "cpf/cnpj", "cnpj fornecedor", "documento fornecedor",
             "cpf", "cnpj cpf"},
    "documento": {"documento", "nota", "nf", "nfe", "nota fiscal", "numero nota",
                  "num nota", "numero da nota", "titulo", "num documento",
                  "numero documento", "n documento", "doc"},
    "valor": {"valor", "valor total", "valor pago", "valor titulo", "valor do titulo",
              "total", "vlr", "montante", "valor liquido", "valor bruto"},
    "data": {"data", "vencimento", "data vencimento", "data de vencimento", "emissao",
             "data emissao", "data de emissao", "data pagamento", "pagamento",
             "competencia", "dt vencimento"},
    "pedido": {"pedido", "oc", "ordem de compra", "num pedido", "numero pedido",
               "pedido compra", "n pedido", "ordem compra"},
}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return " ".join(s.lower().replace(".", " ").replace("_", " ").split())


def detect_mapping(headers: list[str]) -> dict[str, str]:
    """Cabeçalhos crus -> {campo_canonico: cabeçalho_original}. Primeiro match vence."""
    mapping: dict[str, str] = {}
    for raw in headers:
        n = _norm(raw or "")
        for field, syns in _SYNONYMS.items():
            if field in mapping:
                continue
            if n in syns:
                mapping[field] = raw
                break
    return mapping


def parse_amount(v: Any) -> Decimal | None:
    """Aceita '1.234,56', 'R$ 1.234,56', '1234.56', número. Devolve Decimal ou None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if not s:
        return None
    if "," in s:  # formato BR: ponto = milhar, vírgula = decimal
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _read_rows(filename: str, content: bytes) -> tuple[list[str], list[list[Any]]]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook  # import tardio: só quando for XLSX

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:  # CSV (auto-detecta separador , ou ;)
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            delim = dialect.delimiter
        except csv.Error:
            delim = ";" if sample.count(";") >= sample.count(",") else ","
        rows = [list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]
    if not rows:
        return [], []
    headers = ["" if h is None else str(h).strip() for h in rows[0]]
    return headers, rows[1:]


def parse_spreadsheet(filename: str, content: bytes) -> dict[str, Any]:
    """Planilha -> {mapping, headers, rows[]}. Cada row: campos canônicos + _row (1-based)."""
    headers, body = _read_rows(filename, content)
    mapping = detect_mapping(headers)
    idx = {field: headers.index(col) for field, col in mapping.items()}

    def cell(row: list[Any], field: str) -> Any:
        i = idx.get(field)
        if i is None or i >= len(row):
            return None
        v = row[i]
        return v.strip() if isinstance(v, str) else v

    out_rows: list[dict[str, Any]] = []
    for n, row in enumerate(body, start=2):  # linha 1 = cabeçalho
        if row is None or all(c in (None, "") for c in row):
            continue  # linha vazia
        out_rows.append({
            "fornecedor": cell(row, "fornecedor"),
            "cnpj": cell(row, "cnpj"),
            "documento": cell(row, "documento"),
            "valor": cell(row, "valor"),
            "data": cell(row, "data"),
            "pedido": cell(row, "pedido"),
            "_row": n,
        })
    return {"mapping": mapping, "headers": headers, "rows": out_rows}
